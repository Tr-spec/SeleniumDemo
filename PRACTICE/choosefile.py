import time

from selenium import webdriver
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

driver = webdriver.Chrome()
driver.maximize_window()

# upload
# driver.get("https://demo.automationtesting.in/Register.html")
# file = driver.find_element(By.XPATH, "//input[@id='imagesrc']")
# file.send_keys(r"D:\1\upload_pc.jpg")
# time.sleep(10)

# table
# driver.get("https://www.w3schools.com/html/html_tables.asp")
# columns = len(driver.find_elements(By.XPATH, '//*[@id="customers"]/tbody/tr[1]/th'))
# rows = len(driver.find_elements(By.XPATH, '//*[@id="customers"]/tbody/tr'))
#
# print(f"total no. of rows = {rows}")
# print(f"total no. of rows = {columns}")
#
# for r in range(2, rows + 1):
#     for c in range(1, columns + 1):
#         value = driver.find_element(By.XPATH, "//*[@id='customers']/tbody/tr["+str(r)+"]/td["+str(c)+"]").text
#         print(value, end=' ')
#     print()

# Alert
# driver.get("https://demo.automationtesting.in/Alerts.html")
#
# wait = driver.find_element(By.XPATH, "//button[contains(text(),'click the button to display an')]")
# wait.click()
# driver.switch_to.alert.accept()
# time.sleep(5)

# implicitly wait
# driver.get("https://the-internet.herokuapp.com/dynamic_loading/2")
# driver.implicitly_wait(10)  # belongs to selenium
#
# driver.find_element(By.XPATH, "//button[text()='Start']").click()
#
# innerHTML = driver.find_element(By.XPATH, "//h4[normalize-space()='Hello World!']")
# print(innerHTML.get_attribute('innerHTML'))


# explicit wait
# driver.get("https://the-internet.herokuapp.com/dynamic_loading/2")
# driver.find_element(By.XPATH, "//button[text()='Start']").click()
#
# locator = (By.XPATH, "//div[@id='finish']//h4")
# element = WebDriverWait(driver, 10).until(expected_conditions.presence_of_element_located(locator))
# print(element.get_attribute('innerHTML'))

# Excel operations

# path = "D:\ExcelOperation\orders.xlsx"
#
# workbook = openpyxl.load_workbook(path)
#
# sheet = workbook['SSMS']
#
# row = sheet.max_row
# col = sheet.max_column
#
# print(row)
# print(col)
#
# for r in range(1, row + 1):
#     for c in range(1, col + 1):
#         print(sheet.cell(r, c).value, end=' ')
#     print()

# explicit wait
driver.get("https://demo.automationtesting.in/JqueryProgressBar.html")
driver.find_element(By.XPATH, "//button[@id='downloadButton']").click()

locator = (By.XPATH, "(//div[normalize-space()='Complete!'])[2]")
element = WebDriverWait(driver, 30).until(expected_conditions.presence_of_element_located(locator))
print(element.get_attribute('innerHTML'))

time.sleep(3)
driver.find_element(By.XPATH, "//button[normalize-space()='Close']").click()

time.sleep(5)













