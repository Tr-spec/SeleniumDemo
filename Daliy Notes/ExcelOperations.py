import openpyxl

path = "D:\ExcelOperation\python.xlsx"

workbook = openpyxl.load_workbook(path)

sheet = workbook.active

rows = sheet.max_row
cols = sheet.max_column

print(rows)
print(cols)

for r in range(1, rows + 1):
    for c in range(1, cols + 1):
        print(sheet.cell(row=r, column=c).value, end='   ')
    print()


# how to write data in Excel from python
# import openpyxl
#
# path = "D:\ExcelOperation\WirteData.xlsx"
#
# workbook = openpyxl.load_workbook(path)
#
# sheet = workbook.active
#
# for r in range(1, 6):
#     for c in range(1, 4):
#         sheet.cell(row=r, column=c).value = "Welcome"
#
# workbook.save(path)
























